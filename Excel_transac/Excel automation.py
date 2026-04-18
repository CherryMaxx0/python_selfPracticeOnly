import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment


def reduction_system(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    maxi = 8+1
    for row in range(2,maxi):
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.font = Font(size=14)
        corrected_price_cell.value = corrected_price
        corrected_price_cell.number_format = "$#,##0.00"
        corrected_price_cell.alignment = Alignment(horizontal='center',vertical='center')

    sheet.cell(1,4).font = Font(size=14)
    sheet.cell(1,4).alignment = Alignment(horizontal='center',vertical='center')
    sheet.cell(1,4).value = 'Updated Price'
    
    # values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    # chart = BarChart()
    # chart.add_data(values)
    # sheet.add_chart(chart, 'F15')

    wb.save(filename)


# reduction_system('Excel_transac\Transactions.xlsx')